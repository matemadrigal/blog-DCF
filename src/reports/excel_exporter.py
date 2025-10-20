"""Professional Excel Export for DCF Analysis.

This module exports DCF analysis to formatted Excel files with multiple sheets:
- Summary (Resumen Ejecutivo)
- Projections (Proyecciones de FCF)
- Sensitivity (Análisis de Sensibilidad)
- Data (Datos Originales)
"""

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference


class ExcelExporter:
    """Exports DCF analysis to professionally formatted Excel files."""

    def __init__(self):
        """Initialize exporter with default styles."""
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.subheader_font = Font(name='Arial', size=11, bold=True)
        self.subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.normal_font = Font(name='Arial', size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_dcf_analysis(
        self,
        ticker: str,
        fair_value: float,
        current_price: float,
        discount_rate: float,
        growth_rate: float,
        fcf_projections: List[float],
        shares_outstanding: float,
        terminal_value: Optional[float] = None,
        enterprise_value: Optional[float] = None,
        sensitivity_data: Optional[pd.DataFrame] = None,
        scenarios: Optional[Dict[str, Any]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """Export complete DCF analysis to Excel.

        Args:
            ticker: Stock ticker symbol
            fair_value: Calculated fair value per share
            current_price: Current market price
            discount_rate: Discount rate (r)
            growth_rate: Terminal growth rate (g)
            fcf_projections: List of projected FCF values
            shares_outstanding: Number of shares
            terminal_value: Terminal value (optional)
            enterprise_value: Enterprise value (optional)
            sensitivity_data: Sensitivity analysis DataFrame (optional)
            scenarios: Multiple scenarios dict (optional)
            metadata: Additional metadata (optional)

        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb, ticker, fair_value, current_price, discount_rate,
                                   growth_rate, shares_outstanding, enterprise_value,
                                   terminal_value, metadata)

        self._create_projections_sheet(wb, fcf_projections, discount_rate, terminal_value)

        if sensitivity_data is not None:
            self._create_sensitivity_sheet(wb, sensitivity_data)

        if scenarios:
            self._create_scenarios_sheet(wb, scenarios, current_price)

        self._create_data_sheet(wb, ticker, fair_value, current_price, discount_rate,
                               growth_rate, fcf_projections, shares_outstanding, metadata)

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def _create_summary_sheet(self, wb, ticker, fair_value, current_price, discount_rate,
                             growth_rate, shares_outstanding, enterprise_value,
                             terminal_value, metadata):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Resumen Ejecutivo", 0)

        # Title
        ws['A1'] = f'Análisis DCF - {ticker}'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Date
        ws['A2'] = f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws.merge_cells('A2:D2')

        # Key Metrics
        row = 4
        ws[f'A{row}'] = 'MÉTRICAS CLAVE'
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        metrics = [
            ('Fair Value por Acción', f'${fair_value:.2f}'),
            ('Precio de Mercado Actual', f'${current_price:.2f}'),
            ('Upside/Downside', f'{((fair_value - current_price) / current_price * 100):+.2f}%'),
            ('Tasa de Descuento (r)', f'{discount_rate:.2%}'),
            ('Tasa de Crecimiento (g)', f'{growth_rate:.2%}'),
            ('Shares Outstanding', f'{shares_outstanding:,.0f}'),
        ]

        if enterprise_value:
            metrics.append(('Enterprise Value', f'${enterprise_value:,.0f}'))
        if terminal_value:
            metrics.append(('Valor Terminal', f'${terminal_value:,.0f}'))

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.subheader_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.normal_font
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        # Recommendation
        row += 1
        upside = ((fair_value - current_price) / current_price * 100)

        ws[f'A{row}'] = 'RECOMENDACIÓN'
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        if upside > 20:
            recommendation = '🟢 COMPRAR'
            color = '92D050'
        elif upside < -20:
            recommendation = '🔴 VENDER'
            color = 'FF0000'
        else:
            recommendation = '🟡 MANTENER'
            color = 'FFD966'

        ws[f'A{row}'] = recommendation
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row}:D{row}')

        # Metadata if available
        if metadata:
            row += 2
            ws[f'A{row}'] = 'INFORMACIÓN ADICIONAL'
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:D{row}')

            row += 1
            for key, value in metadata.items():
                if key not in ['mode', 'base_fcf', 'growth_rates', 'enhanced_model']:
                    ws[f'A{row}'] = str(key).replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_projections_sheet(self, wb, fcf_projections, discount_rate, terminal_value):
        """Create FCF projections sheet with formulas."""
        ws = wb.create_sheet("Proyecciones FCF")

        # Title
        ws['A1'] = 'Proyecciones de Free Cash Flow'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].fill = self.subheader_fill
        ws.merge_cells('A1:E1')

        # Headers
        headers = ['Año', 'FCF Proyectado', 'Factor de Descuento', 'Valor Presente', 'Acumulado']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows with formulas
        row = 4
        for year, fcf in enumerate(fcf_projections, start=1):
            ws[f'A{row}'] = year
            ws[f'B{row}'] = fcf
            ws[f'B{row}'].number_format = '$#,##0'

            # Formula for discount factor: 1 / (1 + r)^year
            ws[f'C{row}'] = f'=1 / POWER((1 + {discount_rate}), A{row})'
            ws[f'C{row}'].number_format = '0.0000'

            # Formula for present value: FCF * Discount Factor
            ws[f'D{row}'] = f'=B{row} * C{row}'
            ws[f'D{row}'].number_format = '$#,##0'

            # Formula for cumulative
            if row == 4:
                ws[f'E{row}'] = f'=D{row}'
            else:
                ws[f'E{row}'] = f'=E{row-1} + D{row}'
            ws[f'E{row}'].number_format = '$#,##0'

            row += 1

        # Terminal Value row
        if terminal_value:
            row += 1
            ws[f'A{row}'] = 'Valor Terminal'
            ws[f'A{row}'].font = self.subheader_font
            ws[f'B{row}'] = terminal_value
            ws[f'B{row}'].number_format = '$#,##0'

            # Terminal value discount factor
            ws[f'C{row}'] = f'=1 / POWER((1 + {discount_rate}), {len(fcf_projections)})'
            ws[f'C{row}'].number_format = '0.0000'

            # Terminal value PV
            ws[f'D{row}'] = f'=B{row} * C{row}'
            ws[f'D{row}'].number_format = '$#,##0'

            ws[f'E{row}'] = f'=E{row-2} + D{row}'
            ws[f'E{row}'].number_format = '$#,##0'

        # Total Enterprise Value
        row += 2
        ws[f'A{row}'] = 'Enterprise Value Total'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'E{row}'] = f'=E{row-2}'
        ws[f'E{row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'E{row}'].number_format = '$#,##0'

        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

    def _create_sensitivity_sheet(self, wb, sensitivity_df):
        """Create sensitivity analysis sheet."""
        ws = wb.create_sheet("Análisis de Sensibilidad")

        # Title
        ws['A1'] = 'Análisis de Sensibilidad (r vs g)'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].fill = self.subheader_fill
        ws.merge_cells('A1:K1')

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(sensitivity_df, index=True, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                # Format header row
                if r_idx == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center')
                # Format index column
                elif c_idx == 1 and r_idx > 3:
                    cell.font = self.subheader_font
                # Format data cells
                elif c_idx > 1 and r_idx > 3:
                    cell.number_format = '$#,##0.00'

        # Color code based on values
        for row_idx in range(4, ws.max_row + 1):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    # Green for high values, red for low
                    if cell.value > 200:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif cell.value < 100:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 12

    def _create_scenarios_sheet(self, wb, scenarios, current_price):
        """Create enhanced scenarios comparison sheet with risk metrics."""
        ws = wb.create_sheet("Escenarios")

        # Title
        ws['A1'] = 'Análisis de Escenarios (Pesimista/Base/Optimista)'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].fill = self.subheader_fill
        ws.merge_cells('A1:H1')

        # Headers for main table
        headers = ['Escenario', 'Fair Value', 'Upside', 'WACC', 'Terminal Growth', 'Probabilidad', 'Valor Esperado', 'Enterprise Value']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 4
        scenario_names = {'pessimistic': 'Pesimista 🔴', 'base': 'Base 🟡', 'optimistic': 'Optimista 🟢'}
        probabilities = {'pessimistic': 0.25, 'base': 0.50, 'optimistic': 0.25}

        for key, name in scenario_names.items():
            if key in scenarios:
                scenario = scenarios[key]
                fair_value = scenario.get('fair_value_per_share', 0)
                upside = ((fair_value - current_price) / current_price * 100) if current_price > 0 else 0
                wacc = scenario.get('wacc', 0)
                terminal_growth = scenario.get('terminal_growth', 0)
                enterprise_value = scenario.get('enterprise_value', 0)

                ws[f'A{row}'] = name
                ws[f'B{row}'] = fair_value
                ws[f'B{row}'].number_format = '$#,##0.00'
                ws[f'C{row}'] = upside / 100
                ws[f'C{row}'].number_format = '0.00%'
                ws[f'D{row}'] = wacc
                ws[f'D{row}'].number_format = '0.00%'
                ws[f'E{row}'] = terminal_growth
                ws[f'E{row}'].number_format = '0.00%'
                ws[f'F{row}'] = probabilities[key]
                ws[f'F{row}'].number_format = '0%'
                ws[f'G{row}'] = f'=B{row} * F{row}'
                ws[f'G{row}'].number_format = '$#,##0.00'
                ws[f'H{row}'] = enterprise_value
                ws[f'H{row}'].number_format = '$#,##0'

                # Color code upside cells
                if upside > 20:
                    ws[f'C{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif upside < -10:
                    ws[f'C{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    ws[f'C{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

                row += 1

        # Total expected value
        row += 1
        ws[f'A{row}'] = 'Valor Esperado Total (Ponderado)'
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'G{row}'] = f'=SUM(G4:G{row-1})'
        ws[f'G{row}'].font = Font(bold=True, size=12)
        ws[f'G{row}'].number_format = '$#,##0.00'
        ws[f'G{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Add risk metrics section
        row += 3
        ws[f'A{row}'] = 'MÉTRICAS DE RIESGO'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        if 'pessimistic' in scenarios and 'optimistic' in scenarios:
            pess_fv = scenarios['pessimistic'].get('fair_value_per_share', 0)
            opt_fv = scenarios['optimistic'].get('fair_value_per_share', 0)
            base_fv = scenarios['base'].get('fair_value_per_share', 0)

            risk_metrics = [
                ('Rango de Valoración', f'${pess_fv:.2f} - ${opt_fv:.2f}'),
                ('Diferencia del Rango', f'${opt_fv - pess_fv:.2f}'),
                ('Rango %', f'{((opt_fv - pess_fv) / base_fv * 100):.1f}%'),
                ('Precio Actual', f'${current_price:.2f}'),
                ('Downside Risk (Pesimista)', f'{((pess_fv - current_price) / current_price * 100):+.1f}%'),
                ('Upside Potential (Optimista)', f'{((opt_fv - current_price) / current_price * 100):+.1f}%'),
            ]

            for metric, value in risk_metrics:
                ws[f'A{row}'] = metric
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                row += 1

        # Add growth rates comparison
        row += 2
        ws[f'A{row}'] = 'TASAS DE CRECIMIENTO FCF (Promedio por Escenario)'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')

        row += 1
        for key, name in scenario_names.items():
            if key in scenarios:
                scenario = scenarios[key]
                growth_rates = scenario.get('growth_rates', [])
                if growth_rates:
                    avg_growth = sum(growth_rates) / len(growth_rates)
                    ws[f'A{row}'] = name
                    ws[f'B{row}'] = avg_growth
                    ws[f'B{row}'].number_format = '0.00%'
                    ws[f'C{row}'] = ', '.join([f'{g:.1%}' for g in growth_rates])
                    row += 1

        # Column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 17
        ws.column_dimensions['H'].width = 18
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15

    def _create_data_sheet(self, wb, ticker, fair_value, current_price, discount_rate,
                          growth_rate, fcf_projections, shares_outstanding, metadata):
        """Create raw data sheet."""
        ws = wb.create_sheet("Datos Originales")

        # Simple key-value pairs
        data = [
            ('Ticker', ticker),
            ('Fair Value', fair_value),
            ('Current Price', current_price),
            ('Discount Rate (r)', discount_rate),
            ('Growth Rate (g)', growth_rate),
            ('Shares Outstanding', shares_outstanding),
            ('Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]

        for row, (key, value) in enumerate(data, start=1):
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = self.subheader_font
            ws[f'B{row}'] = value

        # FCF Projections
        row = len(data) + 2
        ws[f'A{row}'] = 'FCF Projections'
        ws[f'A{row}'].font = self.subheader_font

        for i, fcf in enumerate(fcf_projections, start=1):
            row += 1
            ws[f'A{row}'] = f'Year {i}'
            ws[f'B{row}'] = fcf

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20


def export_dashboard_to_excel(summary_data: List[Dict[str, Any]]) -> BytesIO:
    """Export dashboard summary to Excel.

    Args:
        summary_data: List of dictionaries with company data

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"

    # Title
    ws['A1'] = 'DCF Portfolio Summary'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ws.merge_cells('A1:H1')

    # Convert to DataFrame
    df = pd.DataFrame(summary_data)

    # Headers
    for col, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # Data
    for r_idx, row in enumerate(df.values, start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value

    # Auto-size columns
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = chr(64 + col_idx)  # A, B, C, ...

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and not isinstance(cell, type(ws.cell(row=1, column=1)).__class__):
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file
